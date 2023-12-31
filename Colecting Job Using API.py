import requests
import pandas as pd
import openpyxl

# Load job data from the API
api_url = "https://cf-courses-data.s3.us.cloud-object-storage.appdomain.cloud/IBM-DA0321EN-SkillsNetwork/labs/module%201/Accessing%20Data%20Using%20APIs/jobs.json"

def get_number_of_jobs_T(technology, job_data):
    number_of_jobs = 0
    for job in job_data:
        key_skills = job.get("Key Skills", "")
        if technology.lower() in key_skills.lower():
            number_of_jobs += 1
    return number_of_jobs

def get_number_of_jobs_L(location, job_data):
    number_of_jobs = 0
    for job in job_data:
        job_location = job.get("Location", "")
        if "us" in job_location.lower() and location.lower() in job_location.lower():
            number_of_jobs += 1
    return number_of_jobs

# Load the job data from the API
response = requests.get(api_url)
if response.status_code == 200:
    job_data = response.json()
else:
    print("Failed to retrieve job data.")
    exit()

# List of technologies and locations
technologies = ["Python", "Java", "SQL", "Machine Learning", "Deep Learning"]
locations = ["Los Angeles", "New York", "Chicago", "San Francisco", "Seattle"]

# Create a DataFrame to store the results
results = []

for tech in technologies:
    tech_jobs = get_number_of_jobs_T(tech, job_data)
    results.append([tech, "Technology", tech_jobs])

for loc in locations:
    loc_jobs = get_number_of_jobs_L(loc, job_data)
    results.append([loc, "Location", loc_jobs])

# Create a DataFrame and save to an Excel file
columns = ["Name", "Type", "Number of Jobs"]
df = pd.DataFrame(results, columns=columns)

excel_filename = "job_results.xlsx"
df.to_excel(excel_filename, index=False)

print(f"Results saved to '{excel_filename}'")

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Rename the worksheet if desired
worksheet.title = "Job Results"

# Now you can add data to the worksheet using worksheet.cell(row, column, value)
worksheet.cell(row=1, column=1, value="Name")
worksheet.cell(row=1, column=2, value="Type")
worksheet.cell(row=1, column=3, value="Number of Jobs")

# Save the workbook
excel_filename = "job_results.xlsx"
workbook.save(excel_filename)

print(f"Workbook '{excel_filename}' created with active worksheet 'Job Results'")

